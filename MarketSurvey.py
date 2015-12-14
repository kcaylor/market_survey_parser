from openpyxl import load_workbook


def _get_kwatcha(value):
    if value[0] == 'K':
        return float(value[1:])
    else:
        return float(value)


class MarketSurvey(object):

    def __init__(self):
        pass

    def init(self, filename=None):
        if filename:
            self.filename = filename
            self.wb = load_workbook(filename=filename)
            self.ws = self.wb.active
            self.crops = []

    def get_district(self):
        # Sometimes the district is write in D9:
        ws = self.ws
        district = None
        district = str(ws.cell('D9').value)
        if district:
            self.district = district
        else:  # Let's try to find it in the actual district field
            district_value = str(ws.cell('C9').value)
            print district_value
            (junk, text) = district_value.split(':')
            # Chop this down to only characters:
            import re
            p = re.compile('[a-z,A-Z]+')
            district = p.findall(text)
            self.district = district

    def get_collector(self):
        # Sometimes the collector is written in O9:
        ws = self.ws
        collector = None
        collector = str(ws.cell('O9').value)
        if collector:
            self.collector = collector
        else:  # Let's try to find it in the actual collector
            collector_value = str(ws.cell('O9').value)
            print collector_value
            (junk, text) = collector_value.split(':')
            # Chop this down to only characters:
            import re
            p = re.compile('[a-z,A-Z]+')
            collector = p.findall(text)
            self.collector = collector

    def get_date(self):
        date = None
        date = self.ws.cell('B9').value
        if date:
            self.date = date

    def get_province(self):
        # Sometimes the district is write in D9:
        province = None
        province = str(self.ws.cell('H9').value)
        if province:
            self.province = province
        else:
            pass

    def is_crop(self, row):
        if self.ws.cell('A' + str(row)).value:
            return True
        else:
            return False

    def parse_row(self, row):
        crop = {}
        crop['code'] = self.ws.cell('A' + str(row)).value
        crop['name'] = self.ws.cell('B' + str(row)).value
        crop['avg_prices'] = []
        if self.ws.cell('I' + str(row)).value:
            crop['avg_prices'].append(_get_kwatcha(
                self.ws.cell('I' + str(row)).value
            ))
        else:
            crop['avg_prices'].append(None)
        if self.ws.cell('Q' + str(row)).value:
            crop['avg_prices'].append(_get_kwatcha(
                self.ws.cell('Q' + str(row)).value
            ))
        else:
            crop['avg_prices'].append(None)
        self.crops.append(crop)
